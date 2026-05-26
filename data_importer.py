"""通用数据导入模块（v22）。

支持 Excel (xlsx/xls) + CSV，自动智能字段映射 + 写入目标表。

工作流：
  1. read_file(path, suffix) → {headers: [...], rows: [...]}
  2. auto_map_fields(headers, target_table) → {excel_col: db_field}（最佳猜测）
  3. import_rows(rows, mapping, target_table, conn) → {success, fail, errors[]}

新加目标表时：
  - 在 TARGETS 注册：display_name / db_table / required_fields / unique_field
  - 在 FIELD_ALIASES[table_key] 加该表的字段中英文别名词典

智能识别策略：
  - 大小写不敏感 + 去标点 + 去空格
  - 别名词典精确匹配（"单位" → unit）
  - 子串匹配（"商品单位" 含 "单位" → unit）
  - 留空：用户在 UI 上手动选择
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import sqlite3
import unicodedata
from pathlib import Path

try:
    import requests
except ImportError:
    requests = None


# ===== DeepSeek LLM 配置 =====

DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"  # V3 模型（speed/cost 优）；改 deepseek-reasoner 用 R1
DEEPSEEK_TIMEOUT = 60  # 单次 LLM 调用最大等待 60 秒


def _load_deepseek_key() -> str | None:
    """优先读环境变量 DEEPSEEK_API_KEY；fallback 读项目根 .deepseek_key 文件。"""
    k = os.environ.get("DEEPSEEK_API_KEY", "").strip()
    if k:
        return k
    key_path = Path(__file__).parent / ".deepseek_key"
    if key_path.exists():
        try:
            return key_path.read_text(encoding="utf-8").strip()
        except OSError:
            return None
    return None


# 支持的目标表
TARGETS = {
    "sku": {
        "label":          "物品 (SKU)",
        "db_table":       "sku",
        "required":       ["name"],
        "unique_field":   "code",
        "auto_gen_code":  True,   # code 留空时自动生成 SP000xxx
        "default_unit":   "份",
    },
    "user": {
        "label":          "用户",
        "db_table":       "user",
        "required":       ["username"],
        "unique_field":   "username",
    },
    "warehouse": {
        "label":          "仓库",
        "db_table":       "warehouse",
        "required":       ["name"],
        "unique_field":   "name",
    },
    "location": {
        "label":          "库位",
        "db_table":       "location",
        "required":       ["warehouse_id", "storage_area", "storage_position"],
        "unique_field":   None,    # 复合唯一键，特殊处理
    },
    "supplier_dropped": None,  # 占位防误用，已 v22 砍掉
}
TARGETS = {k: v for k, v in TARGETS.items() if v is not None}


# 字段别名词典 — key 是 DB 字段名，value 是可能的"用户表头"列表（小写化后匹配）
FIELD_ALIASES = {
    "sku": {
        "code":             ["code", "编号", "编码", "代码", "商品编号", "物品编号", "id", "sku"],
        "name":             ["name", "名称", "物品名称", "商品名称", "产品名称", "title"],
        "spec":             ["spec", "specification", "规格", "规格型号", "型号"],
        "unit":             ["unit", "单位", "计量单位", "度量单位"],
        "safety_stock":     ["safety_stock", "安全库存", "最低库存", "min_stock", "阈值"],
        "cost_price":       ["cost_price", "成本", "成本价", "进价", "采购价"],
        "sale_price":       ["sale_price", "售价", "销售价", "零售价"],
        "brand":            ["brand", "品牌", "牌子"],
        "category":         ["category", "类别", "分类", "物品类别"],
        "category_major":   ["category_major", "大类", "主分类", "物品大类"],
        "usage_purpose":    ["usage_purpose", "用途", "使用用途", "purpose"],
        "image_path":       ["image_path", "图片", "照片", "图片路径", "照片路径", "image", "photo", "picture", "img"],
    },
    "user": {
        "username":         ["username", "用户名", "账号", "login"],
        "display_name":     ["display_name", "name", "姓名", "显示名", "名字"],
        "email":            ["email", "邮箱", "电邮", "e-mail", "mail"],
        "phone":            ["phone", "tel", "手机", "电话", "手机号", "联系电话"],
        "role":             ["role", "角色"],
        "position":         ["position", "岗位", "职位"],
    },
    "warehouse": {
        "name":             ["name", "名称", "仓库名", "仓库名称"],
        "address":          ["address", "地址", "仓库地址"],
        "code":             ["code", "编号", "仓库编号"],
        "note":             ["note", "备注", "说明", "remark"],
    },
    "location": {
        "storage_area":     ["storage_area", "存放区域", "区域", "库区", "area"],
        "storage_position": ["storage_position", "存放位置", "位置", "库位", "库位代码", "库位号", "库位编号", "location_code", "position"],
        "warehouse_id":     ["warehouse_id", "仓库 id", "仓库", "所属仓库"],
        "warehouse_name":   ["warehouse_name", "仓库名", "仓库名称"],
        "note":             ["note", "备注"],
    },
}


def _normalize(s: str) -> str:
    """标准化表头：转 str → 全角转半角 → lower → 去标点 → 去空格"""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.lower().strip()
    s = re.sub(r"[\s\(\)\[\]\{\}（）【】\-_/\\.,;:：。，；！!？\?\*]+", "", s)
    return s


def read_file(file_storage, filename: str) -> dict:
    """读 Excel 或 CSV 文件，返回 {headers: [...], rows: [{header: cell, ...}], errors: [...]}.

    file_storage: werkzeug FileStorage 对象或 BytesIO
    filename: 含扩展名的原始文件名（决定用 openpyxl 还是 csv）
    """
    suffix = Path(filename).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_excel(file_storage)
    if suffix == ".csv":
        return _read_csv(file_storage)
    if suffix == ".xls":
        # 老格式不支持，提示用户另存
        return {"headers": [], "rows": [], "errors": [".xls 格式不支持，请另存为 .xlsx 或导出 CSV"]}
    return {"headers": [], "rows": [], "errors": [f"不支持的文件类型：{suffix}"]}


def _read_excel(file_storage) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"headers": [], "rows": [], "errors": ["缺少 openpyxl 库 (pip install openpyxl)"]}
    try:
        wb = load_workbook(file_storage, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {"headers": [], "rows": [], "errors": ["文件为空"]}
        headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(header_row)]
        rows = []
        for r in rows_iter:
            if all(c is None or str(c).strip() == "" for c in r):
                continue
            row = {}
            for i, val in enumerate(r):
                if i >= len(headers):
                    break
                row[headers[i]] = val if val is not None else ""
            rows.append(row)
        return {"headers": headers, "rows": rows, "errors": []}
    except Exception as e:
        return {"headers": [], "rows": [], "errors": [f"读 Excel 出错：{type(e).__name__}: {e}"]}


def _read_csv(file_storage) -> dict:
    try:
        raw = file_storage.read()
        if isinstance(raw, bytes):
            # 尝试 utf-8-sig（处理 Excel 导出的 BOM）→ utf-8 → gbk
            for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
                try:
                    text = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return {"headers": [], "rows": [], "errors": ["CSV 编码无法识别（试过 utf-8 / gbk）"]}
        else:
            text = raw
        reader = csv.reader(io.StringIO(text))
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            return {"headers": [], "rows": [], "errors": ["CSV 为空"]}
        rows = []
        for r in reader:
            if not r or all((c or "").strip() == "" for c in r):
                continue
            row = {}
            for i, val in enumerate(r):
                if i >= len(headers):
                    break
                row[headers[i]] = val
            rows.append(row)
        return {"headers": headers, "rows": rows, "errors": []}
    except Exception as e:
        return {"headers": [], "rows": [], "errors": [f"读 CSV 出错：{type(e).__name__}: {e}"]}


def auto_map_fields(headers: list, target_table: str) -> dict:
    """智能映射 Excel 表头 → DB 字段名。

    返回 {header: db_field_or_empty}（不能匹配的字段值为 ""，用户在 UI 手动选）
    """
    aliases = FIELD_ALIASES.get(target_table, {})
    if not aliases:
        return {h: "" for h in headers}
    # 把别名表反向展开成 {normalized_alias: db_field}
    rev = {}
    for db_field, alias_list in aliases.items():
        for a in alias_list:
            rev[_normalize(a)] = db_field
    mapping = {}
    used_fields = set()  # 避免同一 DB 字段被多列重复映射
    for h in headers:
        norm = _normalize(h)
        match = ""
        if norm and norm in rev:
            match = rev[norm]
        else:
            # 子串匹配（"商品单位" 含 "单位" → unit）
            for alias_norm, db_field in rev.items():
                if alias_norm and (alias_norm in norm or norm in alias_norm):
                    match = db_field
                    break
        if match and match not in used_fields:
            mapping[h] = match
            used_fields.add(match)
        else:
            mapping[h] = ""
    return mapping


def import_rows(rows: list, mapping: dict, target_table: str, conn: sqlite3.Connection) -> dict:
    """根据映射写入 DB。

    rows: [{header: value, ...}]
    mapping: {header: db_field}（"" 表示该列跳过）
    target_table: TARGETS 的 key
    conn: 已开启的 sqlite3 Connection

    返回 {success: int, fail: int, errors: [str, ...], created_ids: [int, ...]}
    """
    target = TARGETS.get(target_table)
    if not target:
        return {"success": 0, "fail": len(rows), "errors": [f"未知目标表 {target_table}"], "created_ids": []}

    # 反向映射：{db_field: header}（用户可能把多个 header 映射到同一 field，我们取最后一个）
    field_to_header = {v: k for k, v in mapping.items() if v}

    result = {"success": 0, "fail": 0, "errors": [], "created_ids": []}

    # SKU 目标表特殊处理
    if target_table == "sku":
        for i, row in enumerate(rows, start=2):  # 行号从 2 开始（1 是表头）
            try:
                data = _row_to_data(row, field_to_header)
                if not data.get("name") or not str(data.get("name")).strip():
                    raise ValueError("name 必填")
                code = data.get("code") or _next_sku_code(conn)
                conn.execute(
                    "INSERT INTO sku (code, name, spec, unit, safety_stock, cost_price, sale_price, brand, category, category_major, usage_purpose, image_path) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        code,
                        str(data["name"]).strip(),
                        str(data.get("spec") or "").strip() or None,
                        str(data.get("unit") or target.get("default_unit") or "份").strip(),
                        _to_int(data.get("safety_stock"), 0),
                        _to_float(data.get("cost_price"), 0),
                        _to_float(data.get("sale_price"), 0),
                        str(data.get("brand") or "").strip() or None,
                        str(data.get("category") or "").strip() or None,
                        str(data.get("category_major") or "").strip() or None,
                        str(data.get("usage_purpose") or "").strip() or None,
                        str(data.get("image_path") or "").strip() or None,
                    ),
                )
                result["success"] += 1
            except sqlite3.IntegrityError as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：唯一约束冲突（{e}）")
            except Exception as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：{type(e).__name__}: {e}")

    elif target_table == "user":
        from werkzeug.security import generate_password_hash
        default_pw = generate_password_hash("changeme123")
        for i, row in enumerate(rows, start=2):
            try:
                data = _row_to_data(row, field_to_header)
                username = str(data.get("username") or "").strip()
                if not username:
                    raise ValueError("username 必填")
                display_name = str(data.get("display_name") or username).strip()
                conn.execute(
                    "INSERT INTO user (username, password_hash, role, display_name, position, email, phone, must_change_password) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, 1)",
                    (
                        username, default_pw,
                        str(data.get("role") or "staff").strip(),
                        display_name,
                        str(data.get("position") or "").strip() or None,
                        str(data.get("email") or "").strip() or None,
                        str(data.get("phone") or "").strip() or None,
                    ),
                )
                result["success"] += 1
            except sqlite3.IntegrityError as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：用户名重复（{e}）")
            except Exception as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：{type(e).__name__}: {e}")

    elif target_table == "warehouse":
        for i, row in enumerate(rows, start=2):
            try:
                data = _row_to_data(row, field_to_header)
                name = str(data.get("name") or "").strip()
                if not name:
                    raise ValueError("name 必填")
                conn.execute(
                    "INSERT INTO warehouse (name, address, code, note) VALUES (?, ?, ?, ?)",
                    (
                        name,
                        str(data.get("address") or "").strip() or None,
                        str(data.get("code") or "").strip() or None,
                        str(data.get("note") or "").strip() or None,
                    ),
                )
                result["success"] += 1
            except sqlite3.IntegrityError as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：名称重复（{e}）")
            except Exception as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：{type(e).__name__}: {e}")

    elif target_table == "location":
        # 仓库支持按 name 反查（warehouse_name 字段）
        wh_lookup = {r["name"]: r["id"] for r in conn.execute("SELECT id, name FROM warehouse").fetchall()}
        for i, row in enumerate(rows, start=2):
            try:
                data = _row_to_data(row, field_to_header)
                storage_area = str(data.get("storage_area") or "").strip()
                storage_position = str(data.get("storage_position") or "").strip()
                if not storage_area or not storage_position:
                    raise ValueError("存放区域、存放位置都必填")
                wh_id = data.get("warehouse_id")
                if not wh_id and data.get("warehouse_name"):
                    wh_id = wh_lookup.get(str(data["warehouse_name"]).strip())
                wh_id = _to_int(wh_id, 0)
                if not wh_id:
                    raise ValueError("warehouse_id 或 warehouse_name 必须能定位到现有仓库")
                conn.execute(
                    "INSERT INTO location (warehouse_id, storage_area, storage_position, note) VALUES (?, ?, ?, ?)",
                    (
                        wh_id, storage_area, storage_position,
                        str(data.get("note") or "").strip() or None,
                    ),
                )
                result["success"] += 1
            except sqlite3.IntegrityError as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：该仓库已存在相同 存放区域+存放位置（{e}）")
            except Exception as e:
                result["fail"] += 1
                result["errors"].append(f"第 {i} 行：{type(e).__name__}: {e}")

    else:
        result["errors"].append(f"目标表 {target_table} 未实现写入逻辑")

    return result


def _row_to_data(row: dict, field_to_header: dict) -> dict:
    """{header: value} → {db_field: value}（按映射筛掉无关列）"""
    return {field: row.get(header) for field, header in field_to_header.items()}


# ===== LLM 分析（v22 加入 DeepSeek）=====

def llm_analyze_table(headers: list, sample_rows: list, max_samples: int = 5) -> dict:
    """调 DeepSeek 分析整张表 → 返回每列的目标表 + 字段建议 + 多表分流标记。

    返回结构：
    {
        "available": True,                  # LLM 是否可用（key 存在 + 调用成功）
        "columns": [                        # 与 headers 一一对应
            {"header": "...", "target_table": "sku|user|warehouse|location|skip",
             "field": "name|...|''", "confidence": 0.0~1.0, "reason": "..."}
        ],
        "multi_table_split": True/False,    # 是否检测到多表混合
        "summary": "...",                   # 一句话总结（用户预览页顶部显示）
        "warnings": [...],                  # 数据风险提示
        "error": None or "错误描述"
    }

    失败（key 缺失 / 网络问题 / JSON 解析失败）→ available=False，调用方降级到机械识别。
    """
    if requests is None:
        return {"available": False, "error": "requests 库未安装"}
    key = _load_deepseek_key()
    if not key:
        return {"available": False, "error": "未配置 DeepSeek API key（.deepseek_key 或 DEEPSEEK_API_KEY）"}

    # 只取前 max_samples 行做样本（不发太多数据 + 控成本）
    samples = []
    for r in sample_rows[:max_samples]:
        row_dict = {}
        for h in headers:
            v = r.get(h)
            if v is None:
                row_dict[h] = ""
            else:
                s = str(v)
                row_dict[h] = s if len(s) <= 100 else s[:100] + "..."
        samples.append(row_dict)

    # 把目标表 schema 拼成简短描述
    table_schemas = []
    for tk, t in TARGETS.items():
        fields = ", ".join(FIELD_ALIASES.get(tk, {}).keys())
        required = ", ".join(t.get("required", []))
        table_schemas.append(f"  {tk} ({t['label']}) — 必填:[{required}] · 全字段:[{fields}]")
    schema_text = "\n".join(table_schemas)

    prompt = f"""你是仓储系统数据导入助手。下面是用户上传的 Excel 文件的列头和样本数据。请分析每一列应该映射到哪个数据库表的哪个字段。

可用目标表（key + 字段）：
{schema_text}

特殊规则：
- 如果某列无法映射（比如 Excel 自带的 "序号" / "备注栏" 等无 DB 字段对应），target_table 设为 "skip"
- 整张表可能混合了多个目标表的字段（比如同一行既有商品信息也有库位信息）→ multi_table_split = true
- 中文表头要识别同义词（"单位/计量单位/度量单位" → unit；"名称/物品名称/商品名" → name）
- 仔细看样本数据辅助判断（"￥120.00" 大概率是 cost_price 或 sale_price；"1号楼" 是 building）

列头（共 {len(headers)} 列）：
{json.dumps(headers, ensure_ascii=False)}

前 {len(samples)} 行样本：
{json.dumps(samples, ensure_ascii=False, indent=2)}

请只返回 JSON（不要任何解释文字、不要 markdown 代码块）：
{{
  "columns": [
    {{"header": "<列头原文>", "target_table": "sku|user|warehouse|location|skip", "field": "<DB字段名或空>", "confidence": <0-1 浮点>, "reason": "<10 字内中文理由>"}}
  ],
  "multi_table_split": <true|false>,
  "summary": "<一句话总结：例如 '检测到物品 5 列 + 库位 3 列混合'>",
  "warnings": ["<风险或异常>"]
}}
"""

    try:
        resp = requests.post(
            DEEPSEEK_API_URL,
            headers={
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
            },
            json={
                "model": DEEPSEEK_MODEL,
                "messages": [
                    {"role": "system", "content": "你是数据库导入字段映射助手。只返回 JSON，不解释。"},
                    {"role": "user", "content": prompt},
                ],
                "temperature": 0.1,  # 字段映射要稳定
                "response_format": {"type": "json_object"},
                "max_tokens": 2000,
            },
            timeout=DEEPSEEK_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        content = data["choices"][0]["message"]["content"]
        parsed = json.loads(content)
    except requests.exceptions.Timeout:
        return {"available": False, "error": f"DeepSeek API 超时（{DEEPSEEK_TIMEOUT}s）"}
    except requests.exceptions.HTTPError as e:
        return {"available": False, "error": f"DeepSeek API HTTP {e.response.status_code}: {e.response.text[:200]}"}
    except (ValueError, KeyError, json.JSONDecodeError) as e:
        return {"available": False, "error": f"LLM 返回解析失败：{type(e).__name__}: {e}"}
    except Exception as e:
        return {"available": False, "error": f"{type(e).__name__}: {e}"}

    # 规范化返回（防 LLM 返回字段缺失）
    cols = []
    by_header = {c.get("header", ""): c for c in parsed.get("columns", []) if isinstance(c, dict)}
    valid_targets = set(TARGETS.keys()) | {"skip", ""}
    for h in headers:
        c = by_header.get(h, {})
        target = (c.get("target_table") or "skip").strip()
        if target not in valid_targets:
            target = "skip"
        field = (c.get("field") or "").strip()
        # 校验 field 属于该 table 的合法字段
        if target in FIELD_ALIASES and field and field not in FIELD_ALIASES[target]:
            field = ""
        try:
            conf = float(c.get("confidence", 0.5))
        except (ValueError, TypeError):
            conf = 0.5
        cols.append({
            "header": h,
            "target_table": target,
            "field": field,
            "confidence": max(0.0, min(1.0, conf)),
            "reason": str(c.get("reason", ""))[:50],
        })

    return {
        "available": True,
        "columns": cols,
        "multi_table_split": bool(parsed.get("multi_table_split", False)),
        "summary": str(parsed.get("summary", ""))[:200],
        "warnings": [str(w)[:200] for w in (parsed.get("warnings", []) or [])][:5],
        "error": None,
    }


def import_multi_table(rows: list, columns_plan: list, conn: sqlite3.Connection) -> dict:
    """多表分流导入。每个目标表自己 import_rows，合并结果。

    columns_plan: [
        {"header": "...", "target_table": "sku|user|warehouse|location", "field": "code|name|..."},
        ...
    ]
    （target_table == "skip" 或 field == "" 的列自动忽略）

    返回 {
        "tables": {tk: {success, fail, errors, attempted}},
        "total_success": N,
        "total_fail": N,
    }
    """
    # 按 target_table 分组 mapping
    plans_by_table = {}
    for c in columns_plan:
        tk = c.get("target_table")
        field = c.get("field")
        if not tk or tk == "skip" or not field:
            continue
        if tk not in TARGETS:
            continue
        plans_by_table.setdefault(tk, {})[c["header"]] = field

    summary = {"tables": {}, "total_success": 0, "total_fail": 0}
    for tk, mapping in plans_by_table.items():
        # 行内有效性：至少 1 个该表 mapping 的列在该行有非空值才尝试插入
        eligible_rows = []
        for r in rows:
            if any(str(r.get(h, "") or "").strip() for h in mapping):
                eligible_rows.append(r)
        if not eligible_rows:
            summary["tables"][tk] = {"success": 0, "fail": 0, "errors": [], "attempted": 0,
                                     "label": TARGETS[tk]["label"]}
            continue
        result = import_rows(eligible_rows, mapping, tk, conn)
        result["attempted"] = len(eligible_rows)
        result["label"] = TARGETS[tk]["label"]
        summary["tables"][tk] = result
        summary["total_success"] += result.get("success", 0)
        summary["total_fail"] += result.get("fail", 0)
    return summary


def _to_int(v, default=0):
    try:
        if v is None or v == "": return default
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _to_float(v, default=0.0):
    try:
        if v is None or v == "": return default
        return float(v)
    except (ValueError, TypeError):
        return default


def _next_sku_code(conn) -> str:
    """复制 app._next_sku_code 防止循环引用"""
    row = conn.execute(
        "SELECT code FROM sku WHERE code LIKE 'SP%' ORDER BY code DESC LIMIT 1"
    ).fetchone()
    if row and row["code"]:
        try:
            return f"SP{int(row['code'][2:]) + 1:06d}"
        except (ValueError, IndexError):
            pass
    return "SP000001"
