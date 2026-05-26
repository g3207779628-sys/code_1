"""CSV / Excel 导出工具。

调用方约定：
- rows: list of sqlite3.Row / dict / tuple
- columns: list of (header_label, key_or_callable)
- key 是 str 时按 r[key] 取；callable 时按 key(r) 计算
- filename: 不含扩展名
"""
import csv
import io
from typing import Iterable, List, Tuple, Callable, Union

from flask import Response


ColumnSpec = Tuple[str, Union[str, Callable]]


def _value(row, key):
    if callable(key):
        return key(row)
    try:
        v = row[key]
    except (KeyError, IndexError, TypeError):
        return ""
    return "" if v is None else v


def filter_columns(all_cols: List[ColumnSpec], picked: Iterable[str]) -> List[ColumnSpec]:
    """按用户勾选的列标签筛选；空则返回全部。"""
    picked = set(p for p in (picked or []) if p)
    if not picked:
        return all_cols
    out = [c for c in all_cols if c[0] in picked]
    return out or all_cols


def to_csv(rows, columns: List[ColumnSpec], filename: str) -> Response:
    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM，方便 Excel 中文识别
    w = csv.writer(buf)
    w.writerow([c[0] for c in columns])
    for r in rows:
        w.writerow([_value(r, c[1]) for c in columns])
    body = buf.getvalue().encode("utf-8")
    return Response(
        body,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def to_xlsx(rows, columns: List[ColumnSpec], filename: str) -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return to_csv(rows, columns, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = filename[:30] or "Sheet1"
    headers = [c[0] for c in columns]
    ws.append(headers)
    for r in rows:
        ws.append([_value(r, c[1]) for c in columns])
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) + 6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def export(rows, columns: List[ColumnSpec], fmt: str, filename: str) -> Response:
    """统一入口：fmt='csv' 或 'xlsx'（默认 xlsx）。"""
    if (fmt or "").lower() == "csv":
        return to_csv(rows, columns, filename)
    return to_xlsx(rows, columns, filename)
