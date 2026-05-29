# -*- coding: utf-8 -*-
"""导入 file_e 园区仓库清单（84 个真实库位）为权威库位主数据。

- location 表加 6 列：room_name/wh_code/purpose/rectify_status/rectify_due/is_allocated
- 56 个旧派生库位（file_d 来的）整体挪到「待归位」仓库（物品原地不动，保留位置线索）
- 84 个真实库位导入「武汉仓库」，重建 使用部门/分配部门/类型/责任人 字典
- 楼栋号+楼层→storage_area，具体位置描述→storage_position（撞唯一约束自动加后缀）
- 仓库用途/是否分配 从「数据表」视图文件按(编号,位置)补；图片→库位附件
- 物品大类、仓库物资 不导

用法： python import_filee_locations.py --dry-run | --commit
"""
from __future__ import annotations
import sys, re, json, zipfile, argparse
import xml.etree.ElementTree as ET
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")
HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
import database
from openpyxl import load_workbook

MAIN_FILE = Path(r"C:\file_e\(园区仓库清单)表格视图(1).xlsx")
DATA_FILE = Path(r"C:\file_e\(数据表)表格视图.xlsx")
WUHAN = "武汉仓库"
PENDING_WH = "待归位"

NEW_COLS = [
    ("room_name", "TEXT"), ("wh_code", "TEXT"), ("purpose", "TEXT"),
    ("rectify_status", "TEXT"), ("rectify_due", "TEXT"),
    ("is_allocated", "INTEGER DEFAULT 0"),
]

DISPIMG_RE = re.compile(r'DISPIMG\(\s*"([^"]+)"', re.I)
def _ln(t): return t.rsplit("}", 1)[-1]

def build_image_map(path):
    id_to_rid, rid_to_media = {}, {}
    with zipfile.ZipFile(path) as z:
        if "xl/cellimages.xml" not in z.namelist():
            return {}
        root = ET.fromstring(z.read("xl/cellimages.xml"))
        for pic in root.iter():
            if _ln(pic.tag) != "pic": continue
            nm = rid = None
            for el in pic.iter():
                if _ln(el.tag) == "cNvPr": nm = el.get("name")
                elif _ln(el.tag) == "blip":
                    for k, v in el.attrib.items():
                        if _ln(k) == "embed": rid = v
            if nm and rid: id_to_rid[nm] = rid
        rroot = ET.fromstring(z.read("xl/_rels/cellimages.xml.rels"))
        for rel in rroot:
            i, t = rel.get("Id"), rel.get("Target")
            if i and t:
                t = t.replace("\\", "/").lstrip("/")
                rid_to_media[i] = t if t.startswith("xl/") else "xl/" + t
    return {i: rid_to_media[r] for i, r in id_to_rid.items() if r in rid_to_media}

def read_sheet(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it)]
    rows = [r for r in it if any(c is not None and str(c).strip() != "" for c in r)]
    wb.close()
    return headers, rows

def s(v): return str(v).strip() if v is not None else ""
def gi(headers, name):
    return headers.index(name) if name in headers else -1
def cell(row, i): return row[i] if 0 <= i < len(row) and row[i] is not None else ""

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    dry = not args.commit
    print("模式:", "DRY-RUN" if dry else "执行")

    h, rows = read_sheet(MAIN_FILE)
    img_map = build_image_map(MAIN_FILE)
    print("主清单行:", len(rows), "| 图片ID:", len(img_map))

    # 数据表：(编号,位置)->(仓库用途, 是否分配)
    dh, drows = read_sheet(DATA_FILE)
    di_code, di_pos = gi(dh, "仓库编号"), gi(dh, "具体位置描述")
    di_use, di_alloc = gi(dh, "仓库用途"), gi(dh, "是否分配")
    data_lut = {}
    for r in drows:
        key = (s(cell(r, di_code)), s(cell(r, di_pos)))
        data_lut[key] = (s(cell(r, di_use)), s(cell(r, di_alloc)))
    print("数据表补充行:", len(data_lut))

    ix = {n: gi(h, n) for n in ["楼栋号","楼层","具体位置描述","使用部门","仓库名称","分配部门",
          "责任人","预计整改完成时间","备注","仓库编号","类型","整改状态"]}
    photo_cols = [i for i, x in enumerate(h) if "图片" in x or "照片" in x]

    conn = database.get_conn()
    try:
        cur_cols = [r["name"] for r in conn.execute("PRAGMA table_info(location)")]
        # 1. 加列
        for col, typ in NEW_COLS:
            if col not in cur_cols:
                print(f"  ALTER location ADD {col} {typ}")
                if not dry:
                    conn.execute(f"ALTER TABLE location ADD COLUMN {col} {typ}")
        if not dry: conn.commit()

        # 2. 仓库
        def wh_id(name):
            r = conn.execute("SELECT id FROM warehouse WHERE name=?", (name,)).fetchone()
            if r: return r["id"]
            if dry: return -1
            return conn.execute("INSERT INTO warehouse (name) VALUES (?)", (name,)).lastrowid
        wuhan = wh_id(WUHAN)
        pending = wh_id(PENDING_WH)
        print(f"  武汉仓库 id={wuhan}  待归位 id={pending}")

        # 3. 旧库位挪到待归位（保留物品线索）
        old_ids = [r["id"] for r in conn.execute(
            "SELECT id FROM location WHERE warehouse_id=?", (wuhan,)).fetchall()]
        print(f"  现武汉仓库下旧库位 {len(old_ids)} 个 -> 挪到待归位")
        if not dry and old_ids:
            conn.execute(f"UPDATE location SET warehouse_id=? WHERE warehouse_id=?", (pending, wuhan))

        # 4. 字典重建
        for t in ["wh_use_dept", "wh_alloc_dept", "wh_type", "wh_owner"]:
            n = conn.execute(f"SELECT COUNT(*) c FROM {t}").fetchone()["c"]
            print(f"  清字典 {t}: {n} 行")
            if not dry: conn.execute(f"DELETE FROM {t}")
        dict_cache = {}
        def goc(table, name):
            name = s(name)
            if not name: return None
            ck = (table, name)
            if ck in dict_cache: return dict_cache[ck]
            r = conn.execute(f"SELECT id FROM {table} WHERE name=?", (name,)).fetchone()
            if r: dict_cache[ck] = r["id"]; return r["id"]
            if dry: dict_cache[ck] = -len(dict_cache)-1; return dict_cache[ck]
            i = conn.execute(f"INSERT INTO {table} (name) VALUES (?)", (name,)).lastrowid
            dict_cache[ck] = i; return i

        # 5. 导入 84 个真实库位
        stats = {"loc": 0, "img": 0, "purpose": 0, "alloc": 0, "dedup_suffix": 0}
        seen_keys = set()
        zf = None if dry else zipfile.ZipFile(MAIN_FILE)  # 图片源 zip，只开一次
        for r in rows:
            build = s(cell(r, ix["楼栋号"])); floor = s(cell(r, ix["楼层"]))
            area = (build + " " + floor).strip() or "未填"
            pos = s(cell(r, ix["具体位置描述"])) or "未填"
            code = s(cell(r, ix["仓库编号"]))
            # 去重后缀
            key = (area, pos)
            if key in seen_keys:
                stats["dedup_suffix"] += 1
                pos2 = f"{pos}#{code}" if code else pos
                k2 = (area, pos2); n = 1
                while k2 in seen_keys:
                    n += 1; pos2 = f"{pos}#{code}-{n}"; k2 = (area, pos2)
                pos = pos2; key = k2
            seen_keys.add(key)

            use_id = goc("wh_use_dept", cell(r, ix["使用部门"]))
            alloc_id = goc("wh_alloc_dept", cell(r, ix["分配部门"]))
            type_id = goc("wh_type", cell(r, ix["类型"]))
            owner_id = goc("wh_owner", cell(r, ix["责任人"]))
            purpose, alloc_raw = data_lut.get((code, s(cell(r, ix["具体位置描述"]))), ("", ""))
            if purpose: stats["purpose"] += 1
            is_alloc = 1 if "已分配" in alloc_raw else 0
            if is_alloc: stats["alloc"] += 1

            stats["loc"] += 1
            if dry:
                continue
            loc_id = conn.execute(
                "INSERT INTO location (warehouse_id, storage_area, storage_position, "
                "use_dept_id, alloc_dept_id, wh_type_id, resp_owner_id, note, "
                "room_name, wh_code, purpose, rectify_status, rectify_due, is_allocated) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (wuhan, area, pos, use_id, alloc_id, type_id, owner_id,
                 s(cell(r, ix["备注"])) or None, s(cell(r, ix["仓库名称"])) or None,
                 code or None, purpose or None, s(cell(r, ix["整改状态"])) or None,
                 s(cell(r, ix["预计整改完成时间"])) or None, is_alloc)).lastrowid
            # 图片
            paths = []
            for ci in photo_cols:
                m = DISPIMG_RE.search(s(cell(r, ci)))
                if m and m.group(1) in img_map:
                    data = zf.read(img_map[m.group(1)])
                    fname = Path(img_map[m.group(1)]).name
                    tdir = HERE / "uploads" / "库位" / str(loc_id)
                    tdir.mkdir(parents=True, exist_ok=True)
                    (tdir / fname).write_bytes(data)
                    paths.append(f"库位/{loc_id}/{fname}")
                    stats["img"] += 1
            if paths:
                conn.execute("UPDATE location SET attachments=? WHERE id=?",
                             (json.dumps(paths, ensure_ascii=False), loc_id))

        if zf: zf.close()
        if not dry:
            conn.commit()
            print("已提交")
        else:
            conn.rollback()
        print("\n=== 统计 ===")
        for k, v in stats.items(): print(f"  {k}: {v}")
        print(f"  新建字典: 使用部门/分配部门/类型/责任人 共 {len([k for k in dict_cache])} 项")
    finally:
        conn.close()

if __name__ == "__main__":
    main()
