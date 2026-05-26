# -*- coding: utf-8 -*-
"""导入园区物资真实数据（4 张物资清单 + 1 张客服出入库台账）到 warehouse.db。

用法：
    python import_real_data.py --dry-run     # 只模拟，打印将创建什么，不写库、不抽图
    python import_real_data.py --clear        # 真正执行：清空业务数据 + 导入 + 抽图

映射规则（用户确认）：
  物资=物品(sku)；方=单位(owner_unit/owner_party)；
  采购数量=入库单数量(inbound_item.quantity)；采购时间=入库单创建时间(inbound_order.created_at)；
  库存数量/数量=现有库存(inventory.on_hand)；照片=物品图(sku.image_path)。
  客服出入库表：导物品+现库存量；3月/4月/本月(5月)出库 -> stock_log 出库流水（不建出库单）。

保留系统配置表：user / menu_permission / role_preset / notification_channel /
  position / wh_type / wh_alloc_dept / wh_use_dept。
"""
from __future__ import annotations
import sys, os, re, zipfile, argparse, datetime as dt
import xml.etree.ElementTree as ET
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
import database
from openpyxl import load_workbook

FILE_DIR = Path(r"C:\file_d")
INVENTORY_FILES = [
    ("客服", FILE_DIR / "(客服)表格视图(1).xlsx"),
    ("工程", FILE_DIR / "(工程)表格视图.xlsx"),
    ("环境", FILE_DIR / "(环境)表格视图.xlsx"),
    ("秩序", FILE_DIR / "(秩序)表格视图.xlsx"),
]
INOUT_FILE = ("客服出入库", FILE_DIR / "(客服出入库)表格视图.xlsx")

# 保留（系统配置 + 仓库分类字典），其余业务表全清
PRESERVE = {
    "user", "menu_permission", "role_preset", "notification_channel",
    "position", "wh_type", "wh_alloc_dept", "wh_use_dept", "sqlite_sequence",
}

# ---------- WPS DISPIMG 图片解析 ----------
DISPIMG_RE = re.compile(r'DISPIMG\(\s*"([^"]+)"', re.I)

def _ln(tag): return tag.rsplit("}", 1)[-1]

def build_image_map(path):
    """{DISPIMG_ID -> 'xl/media/imageN.ext'}"""
    id_to_rid, rid_to_media = {}, {}
    with zipfile.ZipFile(path) as z:
        if "xl/cellimages.xml" not in z.namelist():
            return {}
        root = ET.fromstring(z.read("xl/cellimages.xml"))
        for pic in root.iter():
            if _ln(pic.tag) != "pic":
                continue
            name_val = rid_val = None
            for el in pic.iter():
                if _ln(el.tag) == "cNvPr":
                    name_val = el.get("name")
                elif _ln(el.tag) == "blip":
                    for k, v in el.attrib.items():
                        if _ln(k) == "embed":
                            rid_val = v
            if name_val and rid_val:
                id_to_rid[name_val] = rid_val
        rroot = ET.fromstring(z.read("xl/_rels/cellimages.xml.rels"))
        for rel in rroot:
            rid, tgt = rel.get("Id"), rel.get("Target")
            if rid and tgt:
                t = tgt.replace("\\", "/").lstrip("/")
                rid_to_media[rid] = t if t.startswith("xl/") else "xl/" + t
    return {i: rid_to_media[r] for i, r in id_to_rid.items() if r in rid_to_media}

def extract_media_bytes(path, media_rel):
    with zipfile.ZipFile(path) as z:
        return z.read(media_rel)

# ---------- 表读取 ----------
def read_rows(path):
    """返回 (headers, [row_tuple,...])，跳过全空行。data_only=False 以拿到 DISPIMG 公式。"""
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it)]
    rows = []
    for r in it:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rows.append(r)
    wb.close()
    return headers, rows

def col_index(headers, *keys, exact=None):
    """按关键字找列索引；exact 指定完全相等优先。返回 -1 找不到。"""
    if exact:
        for i, h in enumerate(headers):
            if h == exact:
                return i
    for key in keys:
        for i, h in enumerate(headers):
            if key in h:
                return i
    return -1

def cell(row, idx):
    if idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else v

def s(v):
    return str(v).strip() if v is not None else ""

def to_int(v, default=0):
    try:
        sv = s(v)
        if sv == "" or sv == "/":
            return default
        return int(float(sv))
    except (ValueError, TypeError):
        return default

def to_float(v, default=0.0):
    try:
        sv = s(v)
        if sv == "" or sv == "/":
            return default
        return float(sv)
    except (ValueError, TypeError):
        return default

def parse_dt(v):
    """采购时间 -> 'YYYY-MM-DD HH:MM:SS' 或 None。"""
    if isinstance(v, (dt.datetime, dt.date)):
        if isinstance(v, dt.date) and not isinstance(v, dt.datetime):
            v = dt.datetime(v.year, v.month, v.day)
        return v.strftime("%Y-%m-%d %H:%M:%S")
    sv = s(v)
    if not sv or sv == "/":
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(sv, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None

def parse_in_contract(v):
    sv = s(v)
    if sv in ("1", "是", "有", "在", "true", "True"):
        return 1
    return 0

# ---------- 主流程 ----------
class Importer:
    def __init__(self, conn, dry):
        self.c = conn
        self.dry = dry
        self.stats = {}
        # 缓存：避免重复 get-or-create
        self._owner_unit = {}
        self._owner_party = {}
        self._cat = {}
        self._cat_major = {}
        self._wh = {}
        self._loc = {}        # (wh_id, area, pos) -> loc_id
        self._sku = {}        # normname -> sku_id
        self._batch = {}      # (sku_id, loc_id) -> batch_id
        self._inv = {}        # (sku_id, loc_id, batch_id) -> True (已建)
        self._order_seq = {}  # prefix -> next seq
        self.admin_id = conn.execute("SELECT id FROM user WHERE username='admin'").fetchone()["id"]

    def bump(self, key, n=1):
        self.stats[key] = self.stats.get(key, 0) + n

    # --- 字典表 get-or-create ---
    def _goc(self, cache, table, name):
        name = s(name)
        if not name:
            return None
        if name in cache:
            return cache[name]
        row = self.c.execute(f"SELECT id FROM {table} WHERE name=?", (name,)).fetchone()
        if row:
            cache[name] = row["id"]; return row["id"]
        if self.dry:
            cache[name] = -len(cache) - 1
            self.bump(f"new_{table}")
            return cache[name]
        cur = self.c.execute(f"INSERT INTO {table} (name) VALUES (?)", (name,))
        cache[name] = cur.lastrowid
        self.bump(f"new_{table}")
        return cur.lastrowid

    def owner_unit(self, n): return self._goc(self._owner_unit, "owner_unit", n)
    def owner_party(self, n): return self._goc(self._owner_party, "owner_party", n)
    def category(self, n): return self._goc(self._cat, "item_category", n)
    def category_major(self, n): return self._goc(self._cat_major, "item_category_major", n)

    def warehouse(self, name):
        name = s(name) or "默认仓"
        if name in self._wh:
            return self._wh[name]
        row = self.c.execute("SELECT id FROM warehouse WHERE name=?", (name,)).fetchone()
        if row:
            self._wh[name] = row["id"]; return row["id"]
        if self.dry:
            self._wh[name] = -len(self._wh) - 1; self.bump("new_warehouse"); return self._wh[name]
        cur = self.c.execute("INSERT INTO warehouse (name) VALUES (?)", (name,))
        self._wh[name] = cur.lastrowid; self.bump("new_warehouse"); return cur.lastrowid

    def location(self, wh_id, area, pos):
        area = s(area) or "默认区"
        pos = s(pos) or "默认位置"
        key = (wh_id, area, pos)
        if key in self._loc:
            return self._loc[key]
        row = self.c.execute(
            "SELECT id FROM location WHERE warehouse_id=? AND storage_area=? AND storage_position=?",
            (wh_id, area, pos)).fetchone()
        if row:
            self._loc[key] = row["id"]; return row["id"]
        if self.dry:
            self._loc[key] = -len(self._loc) - 1; self.bump("new_location"); return self._loc[key]
        cur = self.c.execute(
            "INSERT INTO location (warehouse_id, storage_area, storage_position) VALUES (?,?,?)",
            (wh_id, area, pos))
        self._loc[key] = cur.lastrowid; self.bump("new_location"); return cur.lastrowid

    def _next_order_no(self, prefix):
        if prefix not in self._order_seq:
            today = dt.datetime.now().strftime("%Y%m%d")
            row = self.c.execute(
                "SELECT order_no FROM inbound_order WHERE order_no LIKE ? ORDER BY order_no DESC LIMIT 1",
                (f"{prefix}{today}%",)).fetchone()
            start = int(row["order_no"][-4:]) if row and row["order_no"][-4:].isdigit() else 0
            self._order_seq[prefix] = (today, start)
        today, seq = self._order_seq[prefix]
        seq += 1
        self._order_seq[prefix] = (today, seq)
        return f"{prefix}{today}{seq:04d}"

    @staticmethod
    def normname(name):
        return re.sub(r"\s+", "", s(name)).lower()

    def get_or_create_sku(self, name, **attrs):
        nn = self.normname(name)
        if not nn:
            return None, False
        if nn in self._sku:
            return self._sku[nn], False
        row = self.c.execute("SELECT id FROM sku WHERE name=?", (s(name),)).fetchone()
        if row:
            self._sku[nn] = row["id"]; return row["id"], False
        if self.dry:
            sid = -len(self._sku) - 1; self._sku[nn] = sid; self.bump("new_sku"); return sid, True
        code = self._next_sku_code()
        cur = self.c.execute(
            "INSERT INTO sku (code, name, spec, unit, cost_price, brand, category, category_major, "
            "usage_purpose, in_contract, category_id, category_major_id, owner_unit_id, owner_party_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (code, s(name), attrs.get("spec") or None, attrs.get("unit") or "份",
             attrs.get("cost_price", 0), attrs.get("brand") or None,
             attrs.get("category") or None, attrs.get("category_major") or None,
             attrs.get("usage") or None, attrs.get("in_contract", 0),
             attrs.get("category_id"), attrs.get("category_major_id"),
             attrs.get("owner_unit_id"), attrs.get("owner_party_id")))
        self._sku[nn] = cur.lastrowid; self.bump("new_sku")
        return cur.lastrowid, True

    def _next_sku_code(self):
        row = self.c.execute(
            "SELECT code FROM sku WHERE code LIKE 'SP%' ORDER BY code DESC LIMIT 1").fetchone()
        if row and row["code"]:
            try:
                return f"SP{int(row['code'][2:]) + 1:06d}"
            except (ValueError, IndexError):
                pass
        return "SP000001"

    def first_loc_batch(self, sku_id):
        """返回该 sku 已建的首个 (loc_id, batch_id)；无则 None。"""
        for (sid, loc_id), batch_id in self._batch.items():
            if sid == sku_id:
                return loc_id, batch_id
        return None

    def ensure_inventory(self, sku_id, loc_id, on_hand):
        """保证 (sku,loc) 有 batch + inventory；on_hand 直接设置。返回 (loc_id, batch_id)。"""
        bkey = (sku_id, loc_id)
        if bkey in self._batch:
            batch_id = self._batch[bkey]
        else:
            if self.dry:
                batch_id = -len(self._batch) - 1
            else:
                cur = self.c.execute("INSERT INTO batch (batch_no, sku_id) VALUES (?,?)",
                                     (f"IMP-{sku_id}-{loc_id}", sku_id))
                batch_id = cur.lastrowid
            self._batch[bkey] = batch_id; self.bump("new_batch")
        ikey = (sku_id, loc_id, batch_id)
        if ikey not in self._inv:
            if not self.dry:
                self.c.execute(
                    "INSERT INTO inventory (sku_id, location_id, batch_id, on_hand) VALUES (?,?,?,?)",
                    (sku_id, loc_id, batch_id, max(0, on_hand)))
            self._inv[ikey] = True; self.bump("new_inventory")
        return loc_id, batch_id

    def add_inbound(self, sku_id, loc_id, batch_id, qty, when):
        if qty <= 0:
            return
        self.bump("new_inbound_order"); self.bump("new_inbound_item")
        if self.dry:
            return
        order_no = self._next_order_no("RKD")
        created = when or dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.c.execute(
            "INSERT INTO inbound_order (order_no, status, creator_id, approver_id, note, created_at, approved_at) "
            "VALUES (?, 'approved', ?, ?, '数据导入-历史采购', ?, ?)",
            (order_no, self.admin_id, self.admin_id, created, created))
        self.c.execute(
            "INSERT INTO inbound_item (order_no, sku_id, batch_no, location_id, quantity, unit_price, batch_id) "
            "VALUES (?,?,?,?,?,?,?)",
            (order_no, sku_id, f"IMP-{sku_id}-{loc_id}", loc_id, qty, 0, batch_id))

    def add_outbound_log(self, sku_id, loc_id, batch_id, qty, when, note):
        if qty <= 0:
            return
        self.bump("new_stock_log_out")
        if self.dry:
            return
        self.c.execute(
            "INSERT INTO stock_log (sku_id, batch_id, location_id, delta, source_doc, event_type, occurred_at, operator_id, note) "
            "VALUES (?,?,?,?,?, 'outbound', ?, ?, ?)",
            (sku_id, batch_id, loc_id, -qty, "数据导入-出库台账", when, self.admin_id, note))

    def save_image(self, sku_id, src_path, media_rel):
        """抽 media 字节 -> uploads/物品图/{sku_id}/{name}，更新 sku.image_path。"""
        self.bump("images")
        if self.dry:
            return
        data = extract_media_bytes(src_path, media_rel)
        fname = Path(media_rel).name
        target_dir = HERE / "uploads" / "物品图" / str(sku_id)
        target_dir.mkdir(parents=True, exist_ok=True)
        (target_dir / fname).write_bytes(data)
        rel = f"物品图/{sku_id}/{fname}"
        self.c.execute("UPDATE sku SET image_path=? WHERE id=?", (rel, sku_id))

    # --- 处理一张物资清单 ---
    def import_inventory_sheet(self, dept, path):
        headers, rows = read_rows(path)
        img_map = build_image_map(path)
        ix = {
            "name": col_index(headers, exact="物品名称") if "物品名称" in headers else col_index(headers, "物品名称", "名称"),
            "unit": col_index(headers, exact="单位"),
            "ounit": col_index(headers, "所属单位"),
            "oparty": col_index(headers, "管理方"),
            "cat": col_index(headers, "物品类别", "类别"),
            "catm": col_index(headers, "物品大类", "大类"),
            "spec": col_index(headers, "规格"),
            "brand": col_index(headers, "品牌"),
            "usage": col_index(headers, "用途"),
            "contract": col_index(headers, "合同"),
            "note": col_index(headers, "备注"),
            "price": col_index(headers, "采购单价") if any("采购单价" in h for h in headers) else col_index(headers, "单价"),
            "area": col_index(headers, "存放区域"),
            "pos": col_index(headers, "存放位置"),
            "onhand": col_index(headers, "库存数量") if any("库存数量" in h for h in headers) else col_index(headers, exact="数量"),
            "pqty": col_index(headers, "采购数量") if any("采购数量" in h for h in headers) else col_index(headers, exact="数量"),
            "ptime": col_index(headers, "采购时间"),
        }
        photo_cols = [i for i, h in enumerate(headers) if "照片" in h or "图片" in h]
        wh_id = self.warehouse(f"{dept}仓")
        for r in rows:
            name = s(cell(r, ix["name"]))
            if not name:
                self.bump("skip_noname"); continue
            ou = self.owner_unit(cell(r, ix["ounit"]))
            op = self.owner_party(cell(r, ix["oparty"]))
            cat = self.category(cell(r, ix["cat"]))
            catm = self.category_major(cell(r, ix["catm"]))
            sku_id, created = self.get_or_create_sku(
                name, spec=s(cell(r, ix["spec"])), unit=s(cell(r, ix["unit"])),
                cost_price=to_float(cell(r, ix["price"])), brand=s(cell(r, ix["brand"])),
                category=s(cell(r, ix["cat"])), category_major=s(cell(r, ix["catm"])),
                usage=s(cell(r, ix["usage"])), in_contract=parse_in_contract(cell(r, ix["contract"])),
                category_id=cat, category_major_id=catm, owner_unit_id=ou, owner_party_id=op,
                note=s(cell(r, ix["note"])))
            loc_id = self.location(wh_id, cell(r, ix["area"]), cell(r, ix["pos"]))
            on_hand = to_int(cell(r, ix["onhand"]))
            loc_id, batch_id = self.ensure_inventory(sku_id, loc_id, on_hand)
            self.add_inbound(sku_id, loc_id, batch_id, to_int(cell(r, ix["pqty"])),
                             parse_dt(cell(r, ix["ptime"])))
            # 照片 -> 物品图（仅首图、仅新建 sku 时）
            if created:
                for ci in photo_cols:
                    m = DISPIMG_RE.search(s(cell(r, ci)))
                    if m and m.group(1) in img_map:
                        self.save_image(sku_id, path, img_map[m.group(1)])
                        break
            self.bump(f"rows_{dept}")

    # --- 处理客服出入库台账 ---
    def import_inout_sheet(self, dept, path):
        headers, rows = read_rows(path)
        ix = {
            "name": col_index(headers, "物品名称", "名称"),
            "brand": col_index(headers, "品牌"),
            "spec": col_index(headers, "规格"),
            "unit": col_index(headers, exact="单位"),
            "cur": col_index(headers, "现库存"),
            "m3": col_index(headers, "3月出库"),
            "m4": col_index(headers, "4月出库"),
            "mcur": col_index(headers, "本月出库"),
            "note": col_index(headers, "备注"),
        }
        wh_id = self.warehouse("客服仓")  # 出入库台账属客服部，并入客服仓
        # 出库流水月份（用户：上月=4月，本月=5月；数据为 2025 年）
        month_when = {"m3": "2025-03-31 12:00:00", "m4": "2025-04-30 12:00:00", "mcur": "2025-05-31 12:00:00"}
        month_note = {"m3": "3月出库", "m4": "4月出库", "mcur": "本月(5月)出库"}
        for r in rows:
            name = s(cell(r, ix["name"]))
            if not name:
                self.bump("skip_noname"); continue
            sku_id, _ = self.get_or_create_sku(
                name, spec=s(cell(r, ix["spec"])), unit=s(cell(r, ix["unit"])),
                brand=s(cell(r, ix["brand"])))
            cur_stock = to_int(cell(r, ix["cur"]))
            existing = self.first_loc_batch(sku_id)
            if existing:
                # 已在客服清单出现：复用原库位，不覆盖客服清单的库存数量
                loc_id, batch_id = existing
                self.bump("inout_reuse_loc")
            else:
                loc_id = self.location(wh_id, "默认区", "客服仓库")
                loc_id, batch_id = self.ensure_inventory(sku_id, loc_id, cur_stock)
            for mk in ("m3", "m4", "mcur"):
                qty = to_int(cell(r, ix[mk]))
                self.add_outbound_log(sku_id, loc_id, batch_id, qty, month_when[mk], month_note[mk])
            self.bump(f"rows_{dept}")


def clear_business_data(conn):
    all_tables = [r["name"] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
    targets = [t for t in all_tables if t not in PRESERVE]
    conn.execute("PRAGMA foreign_keys = OFF")
    cleared = {}
    for t in targets:
        try:
            n = conn.execute(f"DELETE FROM \"{t}\"").rowcount
            if n:
                cleared[t] = n
        except Exception as e:
            print(f"  清 {t} 失败: {e}")
    conn.execute("PRAGMA foreign_keys = ON")
    return cleared


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--clear", action="store_true", help="真正执行（清库+导入+抽图）")
    ap.add_argument("--dry-run", action="store_true", help="只模拟")
    args = ap.parse_args()
    dry = not args.clear
    print("模式:", "DRY-RUN（不写库）" if dry else "执行（清库+导入）")

    conn = database.get_conn()
    try:
        if not dry:
            print("\n[1] 清空业务数据（保留系统配置）...")
            cleared = clear_business_data(conn)
            for t, n in sorted(cleared.items(), key=lambda x: -x[1]):
                print(f"    清 {t}: {n}")
            conn.commit()
        else:
            print("\n[1] (dry-run 跳过清库)")

        imp = Importer(conn, dry)
        print("\n[2] 导入 4 张物资清单...")
        for dept, path in INVENTORY_FILES:
            print(f"    -> {dept}  ({path.name})")
            imp.import_inventory_sheet(dept, path)
        print("\n[3] 导入客服出入库台账...")
        d, p = INOUT_FILE
        imp.import_inout_sheet(d, p)

        if not dry:
            conn.commit()
            print("\n已提交。")
        else:
            conn.rollback()

        print("\n===== 统计 =====")
        for k in sorted(imp.stats):
            print(f"  {k}: {imp.stats[k]}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
